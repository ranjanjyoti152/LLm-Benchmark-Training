"""
Excel export functionality for benchmarking results.
Creates structured Excel files with comprehensive analysis.
"""

import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging

from .training import TrainingMetrics

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Handles exporting benchmark results to Excel with charts and formatting.
    """
    
    def __init__(self, output_dir: str = "./benchmark_results"):
        """
        Initialize Excel exporter.
        
        Args:
            output_dir: Directory to save Excel files
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
    def create_metrics_dataframe(self, metrics_history: List[TrainingMetrics]) -> pd.DataFrame:
        """
        Convert metrics history to pandas DataFrame.
        
        Args:
            metrics_history: List of training metrics
            
        Returns:
            DataFrame with all metrics
        """
        data = []
        
        for metric in metrics_history:
            # Base metrics
            row = {
                'timestamp': pd.to_datetime(metric.timestamp, unit='s'),
                'epoch': metric.epoch,
                'step': metric.step,
                'loss': metric.loss,
                'tokens_per_second': metric.tokens_per_second,
                'cpu_utilization': metric.cpu_utilization,
                'ram_utilization': metric.ram_utilization,
                'ram_used_gb': metric.ram_used_gb,
                'ram_total_gb': metric.ram_total_gb,
                'batch_size': metric.batch_size,
                'sequence_length': metric.sequence_length,
                'learning_rate': metric.learning_rate
            }
            
            # GPU metrics - flatten for multiple GPUs
            for gpu_id, utilization in metric.gpu_utilization.items():
                row[f'gpu_{gpu_id}_utilization'] = utilization
            
            for gpu_id, memory_used in metric.gpu_memory_used.items():
                row[f'gpu_{gpu_id}_memory_used_mb'] = memory_used
                
            for gpu_id, memory_total in metric.gpu_memory_total.items():
                row[f'gpu_{gpu_id}_memory_total_mb'] = memory_total
                # Calculate memory utilization percentage
                if memory_total > 0:
                    row[f'gpu_{gpu_id}_memory_utilization'] = (memory_used / memory_total) * 100
                else:
                    row[f'gpu_{gpu_id}_memory_utilization'] = 0
                    
            data.append(row)
        
        return pd.DataFrame(data)
    
    def create_summary_dataframe(self, benchmark_results: Dict[str, Any]) -> pd.DataFrame:
        """
        Create summary DataFrame from benchmark results.
        
        Args:
            benchmark_results: Dictionary containing benchmark results for different models
            
        Returns:
            Summary DataFrame
        """
        summary_data = []
        
        for model_size, results in benchmark_results.items():
            if not results.get('benchmark_completed', False):
                continue
                
            # Get metrics summary
            metrics_summary = results.get('metrics_summary', {})
            
            row = {
                'model_size': model_size,
                'total_training_time_seconds': results.get('total_training_time', 0),
                'total_tokens_processed': results.get('total_tokens_processed', 0),
                'avg_tokens_per_second': results.get('avg_tokens_per_second', 0),
                'final_loss': results.get('final_loss', 0),
                'num_epochs': results.get('num_epochs', 0),
                
                # Performance metrics averages
                'avg_cpu_utilization': metrics_summary.get('cpu_utilization', {}).get('avg', 0),
                'avg_ram_utilization': metrics_summary.get('ram_utilization', {}).get('avg', 0),
                'avg_gpu_utilization': metrics_summary.get('gpu_utilization', {}).get('avg', 0),
                'avg_gpu_memory_utilization': metrics_summary.get('gpu_memory_utilization', {}).get('avg', 0),
                
                # Peak values
                'peak_cpu_utilization': metrics_summary.get('cpu_utilization', {}).get('max', 0),
                'peak_ram_utilization': metrics_summary.get('ram_utilization', {}).get('max', 0),
                'peak_gpu_utilization': metrics_summary.get('gpu_utilization', {}).get('max', 0),
                'peak_gpu_memory_utilization': metrics_summary.get('gpu_memory_utilization', {}).get('max', 0),
                
                # Efficiency metrics
                'tokens_per_second_peak': metrics_summary.get('tokens_per_second', {}).get('max', 0),
                'tokens_per_second_min': metrics_summary.get('tokens_per_second', {}).get('min', 0),
            }
            
            summary_data.append(row)
        
        return pd.DataFrame(summary_data)
    
    def style_worksheet(self, worksheet, title: str):
        """Apply styling to worksheet."""
        # Title
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        worksheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        worksheet['A1'].alignment = Alignment(horizontal="center")
        
        # Header row styling
        if worksheet.max_row > 2:
            for cell in worksheet[3]:  # Assuming headers are in row 3
                if cell.value:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def add_charts(self, workbook, worksheet, df: pd.DataFrame, chart_title_prefix: str):
        """Add charts to worksheet."""
        try:
            # Tokens per second chart
            if 'tokens_per_second' in df.columns:
                chart1 = LineChart()
                chart1.title = f"{chart_title_prefix} - Tokens per Second"
                chart1.y_axis.title = 'Tokens/Second'
                chart1.x_axis.title = 'Step'
                
                # Data for chart
                data = Reference(worksheet, min_col=df.columns.get_loc('tokens_per_second') + 1, 
                               min_row=4, max_row=len(df) + 3)
                cats = Reference(worksheet, min_col=df.columns.get_loc('step') + 1,
                               min_row=4, max_row=len(df) + 3)
                
                chart1.add_data(data, titles_from_data=True)
                chart1.set_categories(cats)
                worksheet.add_chart(chart1, "P2")
            
            # GPU utilization chart
            gpu_util_cols = [col for col in df.columns if 'gpu_' in col and '_utilization' in col and 'memory' not in col]
            if gpu_util_cols:
                chart2 = LineChart()
                chart2.title = f"{chart_title_prefix} - GPU Utilization"
                chart2.y_axis.title = 'Utilization %'
                chart2.x_axis.title = 'Step'
                
                for col in gpu_util_cols[:4]:  # Limit to first 4 GPUs for readability
                    data = Reference(worksheet, min_col=df.columns.get_loc(col) + 1,
                                   min_row=3, max_row=len(df) + 3)
                    chart2.add_data(data, titles_from_data=True)
                
                cats = Reference(worksheet, min_col=df.columns.get_loc('step') + 1,
                               min_row=4, max_row=len(df) + 3)
                chart2.set_categories(cats)
                worksheet.add_chart(chart2, "P18")
            
            # Loss chart
            if 'loss' in df.columns:
                chart3 = LineChart()
                chart3.title = f"{chart_title_prefix} - Training Loss"
                chart3.y_axis.title = 'Loss'
                chart3.x_axis.title = 'Step'
                
                data = Reference(worksheet, min_col=df.columns.get_loc('loss') + 1,
                               min_row=3, max_row=len(df) + 3)
                cats = Reference(worksheet, min_col=df.columns.get_loc('step') + 1,
                               min_row=4, max_row=len(df) + 3)
                
                chart3.add_data(data, titles_from_data=True)
                chart3.set_categories(cats)
                worksheet.add_chart(chart3, "P34")
                
        except Exception as e:
            logger.warning(f"Could not add charts: {e}")
    
    def export_single_model_results(
        self, 
        model_size: str, 
        results: Dict[str, Any], 
        filename: Optional[str] = None
    ) -> str:
        """
        Export results for a single model to Excel.
        
        Args:
            model_size: Size of the model
            results: Benchmark results
            filename: Custom filename (optional)
            
        Returns:
            Path to created Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"benchmark_{model_size}_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Create metrics DataFrame
        metrics_history = results.get('metrics_history', [])
        if not metrics_history:
            logger.warning(f"No metrics history found for {model_size}")
            return filepath
        
        df_metrics = self.create_metrics_dataframe(metrics_history)
        
        # Create workbook
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Raw metrics sheet
            df_metrics.to_excel(writer, sheet_name='Raw_Metrics', index=False, startrow=2)
            
            # Summary sheet
            summary_data = {
                'Metric': [
                    'Model Size',
                    'Total Training Time (s)',
                    'Total Tokens Processed',
                    'Average Tokens/Second',
                    'Final Loss',
                    'Number of Epochs',
                    'Average CPU Utilization (%)',
                    'Average RAM Utilization (%)',
                    'Average GPU Utilization (%)',
                    'Peak Tokens/Second'
                ],
                'Value': [
                    model_size,
                    f"{results.get('total_training_time', 0):.2f}",
                    f"{results.get('total_tokens_processed', 0):,}",
                    f"{results.get('avg_tokens_per_second', 0):.1f}",
                    f"{results.get('final_loss', 0):.4f}",
                    results.get('num_epochs', 0),
                    f"{results.get('metrics_summary', {}).get('cpu_utilization', {}).get('avg', 0):.1f}",
                    f"{results.get('metrics_summary', {}).get('ram_utilization', {}).get('avg', 0):.1f}",
                    f"{results.get('metrics_summary', {}).get('gpu_utilization', {}).get('avg', 0):.1f}",
                    f"{results.get('metrics_summary', {}).get('tokens_per_second', {}).get('max', 0):.1f}"
                ]
            }
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=2)
        
        # Load workbook for styling and charts
        workbook = openpyxl.load_workbook(filepath)
        
        # Style worksheets
        if 'Raw_Metrics' in workbook.sheetnames:
            ws_metrics = workbook['Raw_Metrics']
            self.style_worksheet(ws_metrics, f"Benchmark Results - {model_size} Model")
            self.add_charts(workbook, ws_metrics, df_metrics, f"{model_size} Model")
        
        if 'Summary' in workbook.sheetnames:
            ws_summary = workbook['Summary']
            self.style_worksheet(ws_summary, f"Summary - {model_size} Model")
        
        workbook.save(filepath)
        logger.info(f"Exported {model_size} results to {filepath}")
        
        return filepath
    
    def export_comparison_results(
        self, 
        benchmark_results: Dict[str, Any], 
        filename: Optional[str] = None
    ) -> str:
        """
        Export comparison results across all models to Excel.
        
        Args:
            benchmark_results: Dictionary containing results for all models
            filename: Custom filename (optional)
            
        Returns:
            Path to created Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"benchmark_comparison_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Create summary comparison
        df_summary = self.create_summary_dataframe(benchmark_results)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Summary comparison
            df_summary.to_excel(writer, sheet_name='Model_Comparison', index=False, startrow=2)
            
            # Individual model metrics
            for model_size, results in benchmark_results.items():
                if not results.get('benchmark_completed', False):
                    continue
                    
                metrics_history = results.get('metrics_history', [])
                if metrics_history:
                    df_metrics = self.create_metrics_dataframe(metrics_history)
                    sheet_name = f"{model_size}_Metrics"
                    df_metrics.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        
        # Load workbook for styling and charts
        workbook = openpyxl.load_workbook(filepath)
        
        # Style comparison sheet
        if 'Model_Comparison' in workbook.sheetnames:
            ws_comparison = workbook['Model_Comparison']
            self.style_worksheet(ws_comparison, "Model Performance Comparison")
            
            # Add comparison charts
            try:
                # Tokens per second comparison
                chart = BarChart()
                chart.title = "Average Tokens per Second by Model Size"
                chart.y_axis.title = 'Tokens/Second'
                chart.x_axis.title = 'Model Size'
                
                if 'avg_tokens_per_second' in df_summary.columns:
                    data = Reference(ws_comparison, 
                                   min_col=df_summary.columns.get_loc('avg_tokens_per_second') + 1,
                                   min_row=3, max_row=len(df_summary) + 3)
                    cats = Reference(ws_comparison,
                                   min_col=df_summary.columns.get_loc('model_size') + 1,
                                   min_row=4, max_row=len(df_summary) + 3)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    ws_comparison.add_chart(chart, "M2")
                
            except Exception as e:
                logger.warning(f"Could not add comparison charts: {e}")
        
        # Style individual model sheets
        for model_size, results in benchmark_results.items():
            sheet_name = f"{model_size}_Metrics"
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                self.style_worksheet(ws, f"Detailed Metrics - {model_size} Model")
        
        workbook.save(filepath)
        logger.info(f"Exported comparison results to {filepath}")
        
        return filepath
    
    def export_system_info(self, system_info: Dict[str, Any], filename: Optional[str] = None) -> str:
        """
        Export system information to Excel.
        
        Args:
            system_info: Dictionary containing system information
            filename: Custom filename (optional)
            
        Returns:
            Path to created Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"system_info_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Format system info for Excel
        info_data = []
        
        def flatten_dict(d, parent_key='', sep='_'):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                else:
                    items.append((new_key, v))
            return dict(items)
        
        flat_info = flatten_dict(system_info)
        
        for key, value in flat_info.items():
            info_data.append({'Property': key, 'Value': str(value)})
        
        df_info = pd.DataFrame(info_data)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_info.to_excel(writer, sheet_name='System_Info', index=False, startrow=2)
        
        # Style the worksheet
        workbook = openpyxl.load_workbook(filepath)
        if 'System_Info' in workbook.sheetnames:
            ws = workbook['System_Info']
            self.style_worksheet(ws, "System Information")
        
        workbook.save(filepath)
        logger.info(f"Exported system info to {filepath}")
        
        return filepath